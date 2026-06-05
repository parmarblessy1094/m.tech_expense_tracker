import streamlit as st
import pandas as pd
import io
import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from database.db import (
    get_all_expenses, get_paid_by_totals, get_semester_totals, get_total
)

SEMESTERS = ["Sem 1", "Sem 2", "Sem 3", "Sem 4"]
PRIMARY = "#0F4C81"


def fmt(amount):
    return f"₹{amount:,.2f}"


# ── Excel generation ──────────────────────────────────────────────────────────

def generate_excel(df, paid_totals, sem_totals, grand_total):
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, Fill, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "M.Tech Expenses"

    thin  = Side(style="thin", color="CCCCCC")
    thick = Side(style="medium", color="0F4C81")
    border_thin  = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_thick = Border(left=thick, right=thick, top=thick, bottom=thick)

    def hdr_fill(hex_col):
        return PatternFill("solid", fxFill=hex_col, fgColor=hex_col)

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "M.TECH EXPENSE REPORT"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
    ws["A1"].fill = PatternFill("solid", fgColor="0F4C81")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Generated on: {datetime.now().strftime('%d %B %Y, %I:%M %p')}"
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Column headers row 4
    headers = ["Date", "Semester", "Paid By", "Category", "Description", "Amount (₹)", "Created At"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor="1A5276")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin
    ws.row_dimensions[4].height = 24

    # Data rows
    alt_fill = PatternFill("solid", fgColor="EBF5FB")
    for r_idx, (_, row) in enumerate(df.iterrows(), 5):
        fill = alt_fill if r_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        vals = [row.get("date",""), row.get("semester",""), row.get("paid_by",""),
                row.get("category",""), row.get("description",""),
                row.get("amount", 0), row.get("created_at","")]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill = fill
            cell.border = border_thin
            cell.alignment = Alignment(vertical="center", wrap_text=(c_idx == 5))
            if c_idx == 6:
                cell.number_format = '₹#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")

    last_data_row = 4 + len(df)

    # Semester totals section
    sem_start = last_data_row + 2
    ws.cell(row=sem_start, column=1, value="SEMESTER TOTALS").font = Font(bold=True, size=12, color="0F4C81")
    ws.merge_cells(f"A{sem_start}:G{sem_start}")
    for r_off, (_, sr) in enumerate(sem_totals.iterrows(), 1):
        rc = sem_start + r_off
        ws.cell(row=rc, column=5, value=sr["semester"]).font = Font(bold=True)
        c = ws.cell(row=rc, column=6, value=sr["total"])
        c.number_format = '₹#,##0.00'
        c.alignment = Alignment(horizontal="right")

    # Paid by totals
    pb_start = sem_start + len(sem_totals) + 2
    ws.cell(row=pb_start, column=1, value="PAID BY TOTALS").font = Font(bold=True, size=12, color="0F4C81")
    ws.merge_cells(f"A{pb_start}:G{pb_start}")
    for r_off, (_, pr) in enumerate(paid_totals.iterrows(), 1):
        rc = pb_start + r_off
        ws.cell(row=rc, column=5, value=pr["paid_by"]).font = Font(bold=True)
        c = ws.cell(row=rc, column=6, value=pr["total"])
        c.number_format = '₹#,##0.00'
        c.alignment = Alignment(horizontal="right")

    # Grand total
    gt_row = pb_start + len(paid_totals) + 2
    ws.merge_cells(f"A{gt_row}:E{gt_row}")
    ws.cell(row=gt_row, column=1, value="GRAND TOTAL").font = Font(bold=True, size=13, color="FFFFFF")
    ws.cell(row=gt_row, column=1).fill = PatternFill("solid", fgColor="0F4C81")
    ws.cell(row=gt_row, column=1).alignment = Alignment(horizontal="right", vertical="center")
    gt_cell = ws.cell(row=gt_row, column=6, value=grand_total)
    gt_cell.number_format = '₹#,##0.00'
    gt_cell.font = Font(bold=True, size=13, color="FFFFFF")
    gt_cell.fill = PatternFill("solid", fgColor="0F4C81")
    gt_cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[gt_row].height = 28

    # Column widths
    col_widths = [14, 12, 18, 20, 45, 16, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── PDF generation ────────────────────────────────────────────────────────────

def generate_pdf(df, paid_totals, sem_totals, grand_total):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph,
        Spacer, HRFlowable
    )
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm
    )

    styles = getSampleStyleSheet()
    primary_color  = colors.HexColor("#0F4C81")
    success_color  = colors.HexColor("#2E8B57")
    accent_color   = colors.HexColor("#E8A020")
    light_blue     = colors.HexColor("#EBF5FB")
    mid_blue       = colors.HexColor("#1A5276")
    light_gray     = colors.HexColor("#F5F7FA")
    dark_gray      = colors.HexColor("#333333")

    title_style = ParagraphStyle("title", parent=styles["Title"],
        fontSize=20, textColor=colors.white, spaceAfter=4,
        fontName="Helvetica-Bold", alignment=TA_CENTER)
    subtitle_style = ParagraphStyle("subtitle", parent=styles["Normal"],
        fontSize=11, textColor=colors.HexColor("#BDC3C7"),
        fontName="Helvetica", alignment=TA_CENTER)
    section_style = ParagraphStyle("section", parent=styles["Heading2"],
        fontSize=13, textColor=primary_color, spaceAfter=8, spaceBefore=16,
        fontName="Helvetica-Bold", borderPad=4)
    normal_style = ParagraphStyle("normal_st", parent=styles["Normal"],
        fontSize=9, textColor=dark_gray, fontName="Helvetica")
    footer_style = ParagraphStyle("footer", parent=styles["Normal"],
        fontSize=8, textColor=colors.gray, alignment=TA_CENTER)

    story = []

    # Header banner using table
    header_data = [[
        Paragraph("🎓 M.Tech Expense Report", title_style),
    ]]
    header_table = Table(header_data, colWidths=[doc.width])
    header_table.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, -1), primary_color),
        ("TOPPADDING",  (0, 0), (-1, -1), 16),
        ("BOTTOMPADDING",(0,0), (-1, -1), 16),
        ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("ROUNDEDCORNERS", [8]),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 0.3*cm))

    sub_data = [[
        Paragraph(f"Generated: {datetime.now().strftime('%d %B %Y, %I:%M %p')}", subtitle_style),
        Paragraph(f"Total Records: {len(df)}", subtitle_style),
        Paragraph(f"Grand Total: {fmt(grand_total)}", subtitle_style),
    ]]
    sub_table = Table(sub_data, colWidths=[doc.width/3]*3)
    sub_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), mid_blue),
        ("TEXTCOLOR",  (0,0), (-1,-1), colors.white),
        ("TOPPADDING", (0,0), (-1,-1), 8),
        ("BOTTOMPADDING",(0,0),(-1,-1), 8),
        ("ALIGN",      (0,0), (-1,-1), "CENTER"),
        ("FONTSIZE",   (0,0), (-1,-1), 9),
        ("FONTNAME",   (0,0), (-1,-1), "Helvetica"),
    ]))
    story.append(sub_table)
    story.append(Spacer(1, 0.5*cm))

    # Expense detail table
    story.append(Paragraph("📄 Expense Details", section_style))

    table_header = ["Date", "Semester", "Paid By", "Category", "Description", "Amount (₹)"]
    col_widths = [2.1*cm, 1.8*cm, 3*cm, 3*cm, 5.5*cm, 2.5*cm]

    table_data = [table_header]
    for _, row in df.iterrows():
        table_data.append([
            str(row.get("date", "")),
            str(row.get("semester", "")),
            str(row.get("paid_by", "")),
            str(row.get("category", "")),
            Paragraph(str(row.get("description", "")), ParagraphStyle(
                "desc", parent=normal_style, fontSize=8, leading=10)),
            f"{row.get('amount', 0):,.2f}",
        ])

    main_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    ts = TableStyle([
        # Header
        ("BACKGROUND",  (0, 0), (-1, 0), primary_color),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0), 9),
        ("ALIGN",       (0, 0), (-1, 0), "CENTER"),
        ("TOPPADDING",  (0, 0), (-1, 0), 8),
        ("BOTTOMPADDING",(0, 0),(-1, 0), 8),
        # Data
        ("FONTNAME",    (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",    (0, 1), (-1, -1), 8),
        ("TOPPADDING",  (0, 1), (-1, -1), 5),
        ("BOTTOMPADDING",(0, 1),(-1, -1), 5),
        ("ALIGN",       (5, 1), (5, -1), "RIGHT"),
        ("ALIGN",       (0, 1), (4, -1), "LEFT"),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        # Grid
        ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        # Alternating rows
        *[("BACKGROUND", (0, i), (-1, i), light_blue) for i in range(2, len(table_data), 2)],
        *[("BACKGROUND", (0, i), (-1, i), light_gray) for i in range(1, len(table_data), 2) if i > 0],
    ])
    main_table.setStyle(ts)
    story.append(main_table)
    story.append(Spacer(1, 0.5*cm))

    # Semester totals
    story.append(Paragraph("🎓 Semester-wise Totals", section_style))
    sem_data = [["Semester", "Total Amount"]]
    overall_running = 0
    for _, sr in sem_totals.iterrows():
        sem_data.append([sr["semester"], fmt(sr["total"])])
        overall_running += sr["total"]
    sem_data.append(["GRAND TOTAL", fmt(overall_running)])

    sem_table = Table(sem_data, colWidths=[6*cm, 4*cm])
    sem_table.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0), mid_blue),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME",    (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",    (0, 0), (-1, -1), 10),
        ("ALIGN",       (1, 0), (1, -1), "RIGHT"),
        ("TOPPADDING",  (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING",(0,0), (-1, -1), 7),
        ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("BACKGROUND",  (0, -1), (-1, -1), primary_color),
        ("TEXTCOLOR",   (0, -1), (-1, -1), colors.white),
        ("FONTNAME",    (0, -1), (-1, -1), "Helvetica-Bold"),
        *[("BACKGROUND", (0, i), (-1, i), light_blue) for i in range(2, len(sem_data)-1, 2)],
    ]))
    story.append(sem_table)
    story.append(Spacer(1, 0.5*cm))

    # Paid By totals
    story.append(Paragraph("👥 Paid By Totals", section_style))
    pb_data = [["Paid By", "Total Amount", "% of Total"]]
    for _, pr in paid_totals.iterrows():
        pct = (pr["total"] / grand_total * 100) if grand_total else 0
        pb_data.append([pr["paid_by"], fmt(pr["total"]), f"{pct:.1f}%"])
    pb_data.append(["OVERALL TOTAL", fmt(grand_total), "100.0%"])

    pb_table = Table(pb_data, colWidths=[5*cm, 4*cm, 3*cm])
    pb_table.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0), success_color),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME",    (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",    (0, 0), (-1, -1), 10),
        ("ALIGN",       (1, 0), (-1, -1), "RIGHT"),
        ("TOPPADDING",  (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING",(0,0), (-1, -1), 7),
        ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("BACKGROUND",  (0, -1), (-1, -1), primary_color),
        ("TEXTCOLOR",   (0, -1), (-1, -1), colors.white),
        ("FONTNAME",    (0, -1), (-1, -1), "Helvetica-Bold"),
        *[("BACKGROUND", (0, i), (-1, i), light_blue) for i in range(2, len(pb_data)-1, 2)],
    ]))
    story.append(pb_table)
    story.append(Spacer(1, 0.8*cm))

    # Footer
    story.append(HRFlowable(width="100%", thickness=1, color=primary_color))
    story.append(Spacer(1, 0.2*cm))
    story.append(Paragraph(
        "M.Tech Expense Tracker | Confidential Report | Generated by Expense Management System",
        footer_style
    ))

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


# ── Page render ───────────────────────────────────────────────────────────────

def render():
    st.markdown("## 📁 Reports")
    st.markdown("Download professional Excel and PDF expense reports.")
    st.markdown("---")

    df          = get_all_expenses()
    paid_totals = get_paid_by_totals()
    sem_totals  = get_semester_totals()
    grand_total = get_total()

    if df.empty:
        st.warning("⚠️ No expenses recorded yet. Please add expenses first.")
        return

    # ── Filter for export ─────────────────────────────────────────────────────
    st.markdown("### 🔍 Filter Export Data")
    fc1, fc2 = st.columns(2)
    with fc1:
        f_sem  = st.selectbox("Semester", ["All"] + SEMESTERS, key="rpt_sem")
    with fc2:
        f_paid = st.selectbox("Paid By", ["All"] + get_paid_by_names_list(df), key="rpt_paid")

    filtered = df.copy()
    if f_sem != "All":
        filtered = filtered[filtered["semester"] == f_sem]
    if f_paid != "All":
        filtered = filtered[filtered["paid_by"] == f_paid]

    f_sem_totals  = filtered.groupby("semester")["amount"].sum().reset_index().rename(columns={"amount":"total"})
    f_paid_totals = filtered.groupby("paid_by")["amount"].sum().reset_index().rename(columns={"amount":"total"})
    f_grand       = filtered["amount"].sum()

    st.markdown(f"**{len(filtered)} records** selected · Total: **{fmt(f_grand)}**")
    st.markdown("<br>", unsafe_allow_html=True)

    # ── Download buttons ──────────────────────────────────────────────────────
    st.markdown("### ⬇️ Download Reports")
    rc1, rc2 = st.columns(2)

    with rc1:
        st.markdown("""
        <div class="report-card">
            <div class="rc-icon">📊</div>
            <div class="rc-title">Excel Report</div>
            <div class="rc-desc">Full expense data with formatting, semester totals, paid-by totals, and grand total.</div>
        </div>""", unsafe_allow_html=True)
        if st.button("📊 Generate Excel", use_container_width=True, key="gen_excel", type="primary"):
            with st.spinner("Generating Excel…"):
                excel_bytes = generate_excel(filtered, f_paid_totals, f_sem_totals, f_grand)
            fname = f"MTech_Expenses_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            st.download_button(
                label="⬇️ Download Excel",
                data=excel_bytes,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="dl_excel"
            )
            st.success("✅ Excel report ready!")

    with rc2:
        st.markdown("""
        <div class="report-card">
            <div class="rc-icon">📄</div>
            <div class="rc-title">PDF Report</div>
            <div class="rc-desc">Professional A4 PDF with expense table, semester totals, paid-by breakdown, and summary.</div>
        </div>""", unsafe_allow_html=True)
        if st.button("📄 Generate PDF", use_container_width=True, key="gen_pdf", type="primary"):
            with st.spinner("Generating PDF…"):
                pdf_bytes = generate_pdf(filtered, f_paid_totals, f_sem_totals, f_grand)
            fname = f"MTech_Expenses_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
            st.download_button(
                label="⬇️ Download PDF",
                data=pdf_bytes,
                file_name=fname,
                mime="application/pdf",
                use_container_width=True,
                key="dl_pdf"
            )
            st.success("✅ PDF report ready!")

    # ── Data Preview ──────────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("### 👁️ Data Preview")
    preview = filtered.copy()
    preview["Amount"] = preview["amount"].apply(lambda x: f"₹{x:,.2f}")
    preview = preview[["date","semester","paid_by","category","description","Amount"]].rename(
        columns={"date":"Date","semester":"Semester","paid_by":"Paid By","category":"Category","description":"Description"}
    )
    preview.index = range(1, len(preview)+1)
    st.dataframe(preview, use_container_width=True, height=300)


def get_paid_by_names_list(df):
    return sorted(df["paid_by"].unique().tolist()) if not df.empty else []
